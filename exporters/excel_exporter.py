"""
Excel Exporter — final output layer.

Produces a formatted .xlsx with:
  - Sheet 1: Block C — Fixed Assets
  - Sheet 2: Block D — Working Capital
  - Sheet 3: Audit Log (attempt history)

Cell color coding:
  Green  (#C6EFCE) — verified, high confidence
  Yellow (#FFEB9C) — extracted but low confidence / needs review
  Red    (#FFC7CE) — unverified / formula failure
  White             — zero value (not found in PDF)
"""
from __future__ import annotations

import os
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter

import config
import schemas
from supervisor.orchestrator import AttemptResult, PipelineResult
from utils.logger import get_logger

logger = get_logger("excel_exporter")

# ─────────────────────────────────────────────────────────────────────────────
# Colour palette
# ─────────────────────────────────────────────────────────────────────────────

_GREEN  = PatternFill("solid", fgColor="C6EFCE")
_YELLOW = PatternFill("solid", fgColor="FFEB9C")
_RED    = PatternFill("solid", fgColor="FFC7CE")
_HEADER = PatternFill("solid", fgColor="1F4E79")
_SUBHDR = PatternFill("solid", fgColor="2E75B6")
_TOTAL  = PatternFill("solid", fgColor="D9E1F2")

_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_TOTAL_FONT  = Font(bold=True, size=10)
_NORMAL_FONT = Font(size=9)
_NUM_FORMAT  = '#,##0.00'


def _cell_fill(value: float, conf_dict: Dict, field: str) -> PatternFill:
    if value == 0.0:
        return PatternFill()   # no fill / white
    verified = conf_dict.get(f"{field}_verified", None)
    col_conf  = conf_dict.get(field, 0.0)
    if verified is False:
        return _RED
    if col_conf < config.FUZZY_ACCEPT_THRESHOLD:
        return _YELLOW
    return _GREEN


def _write_header_row(ws, row_num: int, labels: List[str], fill: PatternFill) -> None:
    for col_idx, label in enumerate(labels, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=label)
        cell.fill   = fill
        cell.font   = _HEADER_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws.row_dimensions[row_num].height = 36


def _is_total_row(sl_no: int, block_type: str) -> bool:
    if block_type == "block_c":
        return sl_no in (8, 10)
    return sl_no in (4, 7, 11, 15, 16)


# ─────────────────────────────────────────────────────────────────────────────
# Block C sheet
# ─────────────────────────────────────────────────────────────────────────────

_BLOCK_C_HEADERS = [
    "Sl No", "Types of Asset",
    "Gross: Opening\n01/04", "Gross: Reval\nAddition", "Gross: Actual\nAddition",
    "Gross: Deduction", "Gross: Closing\n31/03",
    "Dep: Up to\nBeginning", "Dep: Provided\nDuring Year",
    "Dep: Adjustment", "Dep: Up to\nEnd",
    "Net: Opening\n01/04", "Net: Closing\n31/03",
]

_BLOCK_C_FIELDS = [
    "gross_opening", "gross_addition_reval", "gross_addition_actual",
    "gross_deduction", "gross_closing",
    "dep_up_to_beginning", "dep_provided_during_year", "dep_adjustment", "dep_up_to_end",
    "net_opening", "net_closing",
]


def _write_block_c(ws, rows: List[Dict[str, Any]]) -> None:
    ws.title = "Block C - Fixed Assets"
    _write_header_row(ws, 1, _BLOCK_C_HEADERS, _HEADER)

    for data_row_idx, row in enumerate(rows, start=2):
        sl   = row.get("sl_no", "")
        conf = row.get("_confidence", {})
        is_total = _is_total_row(sl, "block_c")

        # Sl No
        c = ws.cell(row=data_row_idx, column=1, value=sl)
        c.border = _THIN_BORDER
        c.font   = _TOTAL_FONT if is_total else _NORMAL_FONT
        c.alignment = Alignment(horizontal="center")
        if is_total:
            c.fill = _TOTAL

        # Asset type
        c = ws.cell(row=data_row_idx, column=2, value=row.get("asset_type", ""))
        c.border = _THIN_BORDER
        c.font   = _TOTAL_FONT if is_total else _NORMAL_FONT
        if is_total:
            c.fill = _TOTAL

        # Numeric fields
        for col_offset, field in enumerate(_BLOCK_C_FIELDS, start=3):
            value = float(row.get(field, 0.0) or 0.0) * config.PDF_UNIT_MULTIPLIER
            c = ws.cell(row=data_row_idx, column=col_offset, value=value if value != 0.0 else None)
            c.number_format = _NUM_FORMAT
            c.border  = _THIN_BORDER
            c.font    = _TOTAL_FONT if is_total else _NORMAL_FONT
            c.alignment = Alignment(horizontal="right")
            if is_total:
                c.fill = _TOTAL
            elif value != 0.0:
                c.fill = _cell_fill(value, conf, field)

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    for col in range(3, 14):
        ws.column_dimensions[get_column_letter(col)].width = 14

    ws.freeze_panes = "C2"


# ─────────────────────────────────────────────────────────────────────────────
# Block D sheet
# ─────────────────────────────────────────────────────────────────────────────

_BLOCK_D_HEADERS = ["Sl No", "Item", "Opening (Rs.)", "Closing (Rs.)"]
_BLOCK_D_FIELDS  = ["opening_rs", "closing_rs"]


def _write_block_d(ws, rows: List[Dict[str, Any]]) -> None:
    ws.title = "Block D - Working Capital"
    _write_header_row(ws, 1, _BLOCK_D_HEADERS, _HEADER)

    for data_row_idx, row in enumerate(rows, start=2):
        sl   = row.get("sl_no", "")
        conf = row.get("_confidence", {})
        is_total = _is_total_row(sl, "block_d")

        c = ws.cell(row=data_row_idx, column=1, value=sl)
        c.border = _THIN_BORDER
        c.font   = _TOTAL_FONT if is_total else _NORMAL_FONT
        c.alignment = Alignment(horizontal="center")
        if is_total:
            c.fill = _TOTAL

        c = ws.cell(row=data_row_idx, column=2, value=row.get("item_name", ""))
        c.border = _THIN_BORDER
        c.font   = _TOTAL_FONT if is_total else _NORMAL_FONT
        if is_total:
            c.fill = _TOTAL

        for col_offset, field in enumerate(_BLOCK_D_FIELDS, start=3):
            value = float(row.get(field, 0.0) or 0.0) * config.PDF_UNIT_MULTIPLIER
            c = ws.cell(row=data_row_idx, column=col_offset, value=value if value != 0.0 else None)
            c.number_format = _NUM_FORMAT
            c.border  = _THIN_BORDER
            c.font    = _TOTAL_FONT if is_total else _NORMAL_FONT
            c.alignment = Alignment(horizontal="right")
            if is_total:
                c.fill = _TOTAL
            elif value != 0.0:
                c.fill = _cell_fill(value, conf, field)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.freeze_panes = "C2"


# ─────────────────────────────────────────────────────────────────────────────
# Audit log sheet
# ─────────────────────────────────────────────────────────────────────────────

def _write_audit_log(ws, result: PipelineResult) -> None:
    ws.title = "Audit Log"
    headers = ["Attempt", "Verifier", "Auditor", "Verify Rate", "Elapsed (s)", "Failures"]
    _write_header_row(ws, 1, headers, _SUBHDR)

    for i, att in enumerate(result.attempts, start=2):
        ws.cell(row=i, column=1, value=att.attempt_no)
        ws.cell(row=i, column=2, value=att.verifier_status)
        ws.cell(row=i, column=3, value=att.auditor_status)
        ws.cell(row=i, column=4, value=f"{att.verify_summary.get('rate', 0)*100:.1f}%")
        ws.cell(row=i, column=5, value=att.elapsed_sec)
        ws.cell(row=i, column=6, value="\n".join(att.audit_failures) or "—")
        for col in range(1, 7):
            c = ws.cell(row=i, column=col)
            c.border = _THIN_BORDER
            c.font   = _NORMAL_FONT
            if att.verifier_status == "APPROVED" and att.auditor_status == "APPROVED":
                c.fill = _GREEN
            elif att.verifier_status == "REJECTED" or att.auditor_status == "REJECTED":
                c.fill = _RED
            else:
                c.fill = _YELLOW

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 60

    ws.cell(row=len(result.attempts) + 3, column=1,
            value=f"Total elapsed: {result.total_elapsed}s | Status: {result.final_status}")


# ─────────────────────────────────────────────────────────────────────────────
# Legend sheet
# ─────────────────────────────────────────────────────────────────────────────

def _write_legend(ws) -> None:
    ws.title = "Legend"
    entries = [
        (_GREEN,  "Green  — Value extracted and verified against source text"),
        (_YELLOW, "Yellow — Value extracted but low confidence / needs manual review"),
        (_RED,    "Red    — Value could not be verified against source text"),
        (PatternFill(), "White  — Zero / not found in PDF (may be genuinely absent)"),
        (_TOTAL,  "Blue   — Computed total / sub-total row"),
    ]
    for i, (fill, desc) in enumerate(entries, start=1):
        c = ws.cell(row=i, column=1, value="      ")
        c.fill   = fill
        c.border = _THIN_BORDER
        ws.cell(row=i, column=2, value=desc).font = Font(size=10)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 60


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────

def export(result: PipelineResult, output_path: Optional[str] = None) -> str:
    """
    Export pipeline result to a formatted Excel workbook.

    Args:
        result:      PipelineResult from supervisor
        output_path: full path for .xlsx file; auto-generated if None

    Returns:
        Absolute path to the saved file.
    """
    if output_path is None:
        os.makedirs(config.OUTPUT_DIR, exist_ok=True)
        base = os.path.splitext(os.path.basename(result.pdf_path))[0]
        output_path = os.path.join(config.OUTPUT_DIR, f"{base}_compile_sheet.xlsx")

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    ws_c = wb.create_sheet()
    _write_block_c(ws_c, result.block_c)

    ws_d = wb.create_sheet()
    _write_block_d(ws_d, result.block_d)

    ws_log = wb.create_sheet()
    _write_audit_log(ws_log, result)

    ws_leg = wb.create_sheet()
    _write_legend(ws_leg)

    wb.save(output_path)
    logger.info("Excel saved: %s", output_path)
    return output_path
